#!/usr/bin/env python3
"""
Rebuild SupplementaryTable_SoLD.xlsx from the CSV/TSV tables in table/.

The workbook had drifted badly from the tables on disk (the June 5 build did not
contain four months of regenerated output). Building it with a script means the
workbook is derived from the tables rather than assembled by hand, so it cannot
drift again -- rerun this after any table changes.

Sheet order and names below are the contract. Edit SHEETS to renumber; nothing
else in the script needs touching.

Usage:  python3 scripts/pipeline/build_supplementary_workbook.py [repo_root]
Output: table/supp/SupplementaryTable_SoLD.xlsx
"""
import sys, os, datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"

# (sheet_name, relative_path, description, [optional (column, value) row filter])
SHEETS = [
    ("S1_Ratio_Statistics",        "table/supp/SuppTable_S1_Ratio_Statistics.csv",
     "Class-ratio statistics: effect size, BH-adjusted p, jackknife stability (33 ratios)"),
    ("S2_Species_Jackknife",       "table/supp/SuppTable_S2_Species_Jackknife.csv",
     "Species-level CLR contrast, BH-adjusted p and jackknife stability (152 shared species)"),
    ("S3_Species_Stability_by_class","table/supp/SuppTable_S3_Species_Stability_by_Class.csv",
     "Per-class rollup of species significance and stability"),
    ("S4_TopVariance_Lipids",      "table/supp/SuppTable_S4_TopVariance_Lipids.csv",
     "Top 10 highest-variance lipid species per trial"),
    ("S5A_Class_CLR_Contrast",     "table/supp/SuppTable_S5A_Class_CLR_Contrast.csv",
     "Class-level CLR contrast (LIN - CTL) with confidence intervals"),
    ("S5B_Class_ALR_Contrast",     "table/supp/SuppTable_S5B_Class_ALR_Contrast.csv",
     "Class-level ALR contrast, TG-referenced"),
    ("S5C_Class_CLR_Corr_Delta",   "table/supp/SuppTable_S5C_Class_CLR_Correlation_Delta.csv",
     "Class-pair CLR correlations in each trial, delta r, and sign-reversal flag"),
    ("S6a_Species_Summary",        "table/supp/SuppTable_S6a_Species_Summary.csv",
     "Species counts per trial, shared and trial-specific"),
    ("S6b_Species_by_Class",       "table/supp/SuppTable_S6b_Species_by_Class.csv",
     "Species counts by lipid class"),
    ("S6c_Species_by_SuperClass",  "table/supp/SuppTable_S6c_Species_by_SuperClass.csv",
     "Species counts by lipid superclass"),
    ("S7_CTL_GWAS_candidates_ind", "table/supp/SuppTable_S7_CTL_GWAS_candidate_genes_for_individual_lipid_traits.tsv",
     "CTL candidate genes, individual lipid traits (LD r2 >= 0.4)"),
    ("S8_CTL_GWAS_candidates_sumratio","table/supp/SuppTable_S8_CTL_GWAS_candidate_genes_for_lipid_sum_ratio_traits.tsv",
     "CTL candidate genes, class sums and ratios"),
    ("S9_LIN_GWAS_candidates_ind", "table/supp/SuppTable_S9_LIN_GWAS_candidate_genes_for_individual_lipid_traits.tsv",
     "LIN candidate genes, individual lipid traits"),
    ("S10_LIN_GWAS_candidate_sumratio","table/supp/SuppTable_S10_LIN_GWAS_candidate_genes_for_lipid_sum_ratio_traits.tsv",
     "LIN candidate genes, class sums and ratios"),
    ("S11_LINEX_Reactions",        "table/supp/SuppTable_S11_LINEX_Reactions.csv",
     "Candidate reaction branches, substrates, products and RHEA identifiers"),
    ("S12_ReactionBalance",        "table/supp/SuppTable_S12_ReactionBalance.csv",
     "Reaction-balance scores, LIN - CTL, paired genotype-matched Wilcoxon"),
    ("S13_LINEX_GWAS_GeneSupport", "table/supp/SuppTable_S13_LINEX_GWAS_GeneSupport.csv",
     "GWAS candidate support for each reaction branch"),
    ("S14_LINEX_GWAS_BranchSummary","table/supp/SuppTable_S14_LINEX_GWAS_BranchSummary.csv",
     "Branch-level summary of GWAS support"),
    ("S15_final_lipid_classes",    "table/supp/SuppTable_S15_final_lipid_classes.csv",
     "Annotation table mapping every feature to class, subclass and superclass"),
    ("S16_Genomic_Heritability",   "table/new_table/SuppTable_Genomic_Heritability_All.csv",
     "Genomic heritability per feature and per class sum, both trials"),
    ("S17_GO_BP_all_terms",        "table/go_enrichment/Table_GO_enrichment_all.tsv",
     "GO biological-process enrichment, all terms, with gene-level and LD-aware q",
     ("Ontology", "BP")),
    ("S18_GO_MF_all_terms",        "table/go_enrichment/Table_GO_enrichment_all.tsv",
     "GO molecular-function enrichment, all terms, with gene-level and LD-aware q",
     ("Ontology", "MF")),
    ("S19_Genes_in_enriched_terms","table/go_enrichment/genes_in_enriched_GO_terms.tsv",
     "Candidate genes carrying each enriched GO term"),
    # S20 intentionally unused. The recurrence-selected GO enrichment
    # (curated_gwas_themes/go_enrichment/*_by_recurrence.tsv) was cited here until a
    # reviewer asked for that sentence to be cut, so the analysis is no longer
    # referenced by the manuscript. The number is left as a gap rather than
    # renumbering S21-S27, which would break the "S21 to S24" citation.
    ("S21_Overlap_gene_level",     "table/overlap/gwas_overlap_overall.csv",
     "CTL/LIN candidate-gene overlap, gene level, per trait layer"),
    ("S22_Overlap_locus_level",    "table/overlap/gwas_overlap_locus_level.csv",
     "CTL/LIN overlap collapsed to genomic windows at 100, 250 and 500 kb"),
    ("S23_Shared_genes_individual","table/overlap/shared_genes_individual.csv",
     "Genes shared between trials, individual-lipid layer"),
    ("S24_Shared_genes_sumratio",  "table/overlap/shared_genes_sumratio.csv",
     "Genes shared between trials, class-sum/ratio layer"),
    ("S25_Race_Structure_Tests",   "table/race/race_structure_lipid_tests.csv",
     "Kruskal-Wallis tests of class sums by botanical race and by genome-wide cluster"),
    ("S26_OPLS_VIP_ratios",        "table/supp/SuppTable_S7_OPLS_VIP_ratios.csv",
     "OPLS-DA VIP scores for class ratios (filename still says S7; renumbered here)"),
    ("S27_SERRF_Traceability",     "table/new_table/SuppTable_SERRF_Scan_to_Lipid_Mapping_Audit.csv",
     "SERRF scan-to-lipid mapping audit"),
]

def main(root):
    out = os.path.join(root, "table/supp/SupplementaryTable_SoLD.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)

    index_rows, frames, missing = [], [], []
    for entry in SHEETS:
        name, rel, desc = entry[0], entry[1], entry[2]
        row_filter = entry[3] if len(entry) > 3 else None
        if isinstance(rel, list):          # several files -> one sheet
            parts, srcs = [], []
            for sub, label in rel:
                sp = os.path.join(root, sub)
                if not os.path.exists(sp):
                    missing.append(sub); continue
                d = pd.read_csv(sp, sep="\t" if sp.endswith(".tsv") else ",", low_memory=False)
                d.insert(0, "Condition", label)
                parts.append(d); srcs.append(sub)
            if not parts: continue
            df = pd.concat(parts, ignore_index=True)
            path = os.path.join(root, srcs[0]); rel = " + ".join(srcs)
        else:
            path = os.path.join(root, rel)
            if not os.path.exists(path):
                missing.append(rel); continue
            sep = "\t" if path.endswith(".tsv") else ","
            df = pd.read_csv(path, sep=sep, low_memory=False)
        if row_filter:
            col, val = row_filter
            df = df[df[col] == val].reset_index(drop=True)
        frames.append((name, df))
        stamp = datetime.datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d")
        index_rows.append({"Sheet": name, "Description": desc, "Rows": len(df),
                           "Columns": df.shape[1], "Source file": rel, "Source last modified": stamp})

    index = pd.DataFrame(index_rows)

    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        index.to_excel(xw, sheet_name="Index", index=False)
        for name, df in frames:
            df.to_excel(xw, sheet_name=name[:31], index=False)

    # formatting
    wb = load_workbook(out)
    head_fill = PatternFill("solid", fgColor="DDDDDD")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.font = Font(name=FONT, bold=True); c.fill = head_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = max((len(str(ws.cell(r, col).value)) for r in range(1, min(ws.max_row, 200) + 1)
                         if ws.cell(r, col).value is not None), default=10)
            ws.column_dimensions[letter].width = min(max(width + 2, 10), 60)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = Font(name=FONT)
    wb.save(out)

    print(f"wrote {out}")
    print(f"  {len(frames)} data sheets + Index")
    for name, df in frames:
        print(f"    {name:34} {len(df):>7} rows")
    if missing:
        print("\n  MISSING (skipped):")
        for m in missing: print("   ", m)

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else ".")
